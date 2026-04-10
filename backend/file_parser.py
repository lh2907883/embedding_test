import io
from pathlib import Path


def extract_text(filename: str, content: bytes) -> str:
    suffix = Path(filename).suffix.lower()
    extractors = {
        ".pdf": _extract_pdf,
        ".docx": _extract_docx,
        ".xlsx": _extract_excel,
        ".txt": _extract_txt,
    }
    extractor = extractors.get(suffix)
    if extractor is None:
        raise ValueError(f"不支持的文件类型: {suffix}，仅支持 .pdf/.docx/.xlsx/.txt")
    return extractor(content)


def _extract_pdf(content: bytes) -> str:
    from PyPDF2 import PdfReader
    reader = PdfReader(io.BytesIO(content))
    return "\n".join(page.extract_text() or "" for page in reader.pages)


def _extract_docx(content: bytes) -> str:
    from docx import Document
    doc = Document(io.BytesIO(content))
    return "\n".join(para.text for para in doc.paragraphs)


def _extract_excel(content: bytes) -> str:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    lines = []
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            line = "\t".join(str(cell) if cell is not None else "" for cell in row)
            if line.strip():
                lines.append(line)
    return "\n".join(lines)


def _extract_txt(content: bytes) -> str:
    for encoding in ("utf-8", "gbk", "latin-1"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    return content.decode("utf-8", errors="replace")
